"""
routes/pricebook.py — Pricebook CRUD endpoints for CotizaExpress.

Upload, list, create, update, delete pricebook items.
Synonym generation, deduplication, rebuild-synonyms.
"""

import json
import logging
import re
import traceback
from io import BytesIO
from typing import Optional

from fastapi import (
    APIRouter,
    BackgroundTasks,
    File,
    Header,
    HTTPException,
    Query,
    Request,
    UploadFile,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from psycopg2 import IntegrityError
from pydantic import BaseModel

from auth import get_company_from_bearer, get_user_from_session, require_company_id
from db import get_conn
from semantic_search import (
    auto_generate_context_groups,
    rebuild_embeddings_for_company,
    upsert_single_embedding,
)

log = logging.getLogger("cotizaexpress.pricebook")

router = APIRouter()


# ── Pydantic models ────────────────────────────────────────────────────────

class PricebookItemCreateBody(BaseModel):
    name: str
    sku: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    vat_rate: Optional[float] = 0.16
    source: Optional[str] = "manual"
    synonyms: Optional[str] = None
    bundle_size: Optional[int] = None


class PricebookItemUpdateBody(BaseModel):
    name: Optional[str] = None
    sku: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    vat_rate: Optional[float] = None
    synonyms: Optional[str] = None
    bundle_size: Optional[int] = None
    is_default: Optional[bool] = None


class PricebookBulkBody(BaseModel):
    items: list  # list of {name, price, unit, category}


# ── Helpers (imported lazily from server to avoid circular imports) ──────

def _get_openai_client():
    """Lazy import of the openai_client global from server."""
    from server import openai_client
    return openai_client


def _norm_name(s: str) -> str:
    """Normalize product name for dedup key."""
    return " ".join((s or "").strip().lower().split())


def _normalize_display_name(name: str) -> str:
    """Primera letra mayúscula, resto minúsculas."""
    n = (name or "").strip()
    return (n[0].upper() + n[1:].lower()) if n else n


def _auto_plural_singular(name: str) -> list:
    """Disabled: auto-plurals cause more harm than good."""
    return []


def _rebuild_embeddings_bg(company_id: str):
    """Background task to rebuild embeddings + context groups + LLM context."""
    try:
        log.info("BG EMBEDDINGS START: company=%s", company_id)
        conn = get_conn()
        try:
            result = rebuild_embeddings_for_company(conn, company_id)
            log.info("BG EMBEDDINGS DONE: %s", result)
            try:
                cg_result = auto_generate_context_groups(conn, company_id)
                log.debug("BG CONTEXT GROUPS: %s", cg_result.get("status"))
            except Exception as cge:
                log.error("BG CONTEXT GROUPS ERROR: %s", repr(cge))
        finally:
            conn.close()

        # Regenerate LLM context (jerga hints + system prompt) based on new catalog
        try:
            from llm_context_generator import generate_and_store_llm_context
            generate_and_store_llm_context(company_id)
        except Exception as lce:
            log.error("BG LLM CONTEXT ERROR: %s", repr(lce))

    except Exception as e:
        log.error("BG EMBEDDINGS ERROR: %s", repr(e))


# ── Routes ──────────────────────────────────────────────────────────────────

@router.post("/api/pricebook/rebuild-synonyms")
def rebuild_synonyms(request: Request):
    company_id = require_company_id(request)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, name, synonyms FROM pricebook_items WHERE company_id=%s",
            (company_id,)
        )
        rows = cur.fetchall()
        updated = 0
        for item_id, name, synonyms in rows:
            existing = (synonyms or "").strip()
            auto_vars = _auto_plural_singular(name)
            if auto_vars:
                existing_set = {s.strip().lower() for s in existing.split(",") if s.strip()}
                new_vars = [v for v in auto_vars if v not in existing_set]
                if new_vars:
                    new_synonyms = (existing + ", " + ", ".join(new_vars)).strip(", ")
                    cur.execute(
                        "UPDATE pricebook_items SET synonyms=%s, updated_at=now() WHERE id=%s",
                        (new_synonyms, item_id)
                    )
                    updated += 1
        return {"ok": True, "company_id": company_id, "total": len(rows), "updated": updated}
    finally:
        cur.close()
        conn.close()


@router.get("/api/web/pricebook/template")
@router.get("/api/carga-productos/template")  # alias for CargaProductos.js
def download_template(request: Request):
    _ = get_user_from_session(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "pricebook"
    ws.append(["nombre", "precio_base", "unidad", "sku", "vat_rate"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = "plantilla-cotizaexpress.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/carga-productos/rapida")
async def carga_productos_rapida(request: Request, background_tasks: BackgroundTasks = None):
    """Carga rápida de productos: recibe lista de {nombre, precio_base, categoria, unidad}."""
    user = get_user_from_session(request)
    company_id = require_company_id(request)
    body = await request.json()
    productos = body.get("productos") or []
    if not productos:
        raise HTTPException(status_code=400, detail="No hay productos para cargar")

    openai_client = _get_openai_client()
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        insertados = 0
        actualizados = 0
        errores = []

        # Generar sinónimos por batch
        names_list = [_normalize_display_name(p["nombre"]) for p in productos if p.get("nombre")]
        synonyms_map = {}
        if openai_client:
            for i in range(0, len(names_list), 20):
                batch = names_list[i:i+20]
                try:
                    numbered = "\n".join(f"{j+1}. {n}" for j, n in enumerate(batch))
                    resp = openai_client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {"role": "system", "content": (
                                "Eres experto en ferreterías de México. Para cada producto numerado, "
                                "genera hasta 5 sinónimos coloquiales, typos comunes o marcas usadas como nombre genérico. "
                                'Responde SOLO en JSON válido así: {"1": "sin1, sin2", "2": "sin1, sin2"} '
                                "Sin explicación, sin markdown, solo el JSON."
                            )},
                            {"role": "user", "content": numbered}
                        ],
                        temperature=0.3, max_tokens=300,
                    )
                    raw = (resp.choices[0].message.content or "{}").strip().replace("```json", "").replace("```", "").strip()
                    parsed_syns = json.loads(raw)
                    for k, v in parsed_syns.items():
                        if k.isdigit() and int(k)-1 < len(batch):
                            synonyms_map[batch[int(k)-1]] = v
                except Exception as e:
                    log.error("BATCH SYNONYMS ERROR (rapida): %s", repr(e))

        for p in productos:
            nombre = _normalize_display_name(p.get("nombre", "").strip())
            if not nombre:
                continue
            try:
                precio = float(str(p.get("precio_base", 0)).replace("$", "").replace(",", ""))
            except Exception:
                errores.append(f"Precio inválido para '{nombre}'")
                continue
            if precio <= 0:
                errores.append(f"Precio debe ser > 0 para '{nombre}'")
                continue

            unidad = p.get("unidad", "Pieza") or "Pieza"
            name_norm = _norm_name(nombre)
            auto_syn = synonyms_map.get(nombre, "")

            try:
                cur.execute(
                    """
                    INSERT INTO pricebook_items
                        (company_id, name, name_norm, unit, price, synonyms, source, updated_at)
                    VALUES
                        (%s, %s, %s, %s, %s, %s, 'rapida', now())
                    ON CONFLICT (company_id, name_norm)
                    DO UPDATE SET
                        name = EXCLUDED.name, unit = EXCLUDED.unit,
                        price = EXCLUDED.price,
                        synonyms = COALESCE(NULLIF(pricebook_items.synonyms, ''), EXCLUDED.synonyms),
                        source = 'rapida', updated_at = now()
                    RETURNING (xmax = 0) AS is_insert
                    """,
                    (company_id, nombre, name_norm, unidad, precio, auto_syn),
                )
                row = cur.fetchone()
                if row and row[0]:
                    insertados += 1
                else:
                    actualizados += 1
            except Exception as e:
                log.error("RAPIDA INSERT ERROR %s: %s", nombre, repr(e))
                errores.append(f"Error con '{nombre}': {str(e)}")

        conn.commit()

        # Rebuild embeddings + context groups en background
        if background_tasks:
            background_tasks.add_task(_rebuild_embeddings_bg, company_id)
        else:
            try:
                rebuild_embeddings_for_company(conn, company_id)
                auto_generate_context_groups(conn, company_id)
            except Exception as e:
                log.error("EMBEDDINGS/CTX REBUILD ERROR (rapida): %s", repr(e))

        return {
            "ok": True,
            "productos_insertados": insertados,
            "productos_actualizados": actualizados,
            "errores": errores,
        }
    except HTTPException:
        raise
    except Exception as e:
        log.error("CARGA RAPIDA ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.post("/api/pricebook/upload")
@router.post("/api/carga-productos/upload-excel")  # alias for CargaProductos.js
def pricebook_upload(
    request: Request,
    authorization: str = Header(default=""),
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
):
    if authorization and authorization.lower().startswith("bearer "):
        tenant = get_company_from_bearer(authorization)
        company_id = tenant["company_id"]
    else:
        company_id = require_company_id(request)

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Solo archivos .xlsx o .xlsm")

    openai_client = _get_openai_client()
    conn = None
    cur = None
    upload_id = None

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO pricebook_uploads (company_id, filename, status) VALUES (%s, %s, 'processing') RETURNING id",
            (company_id, file.filename),
        )
        upload_id = cur.fetchone()[0]

        content = file.file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers_raw = [str(h or "") for h in header_row]
        headers_norm = [h.strip().lower() for h in headers_raw]

        alias = {
            "nombre": "name", "producto": "name", "product": "name",
            "precio": "price", "precio_base": "price", "precio unitario": "price",
            "costo": "price", "cost": "price",
            "unidad": "unit", "uom": "unit",
            "vat_rate": "vat_rate", "iva": "vat_rate",
            "sku": "sku",
        }

        headers_mapped = [alias.get(h, h) for h in headers_norm]
        idx = {h: i for i, h in enumerate(headers_mapped)}

        required = {"name", "price"}
        missing_cols = required - set(headers_mapped)
        if missing_cols:
            raise HTTPException(
                status_code=400,
                detail={"error": f"Faltan columnas requeridas: {sorted(missing_cols)}", "headers_detectadas": headers_norm, "headers_mapeadas": headers_mapped},
            )

        parsed_rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None or str(v).strip() == "" for v in r):
                continue
            name = _normalize_display_name(str(r[idx["name"]])) if r[idx["name"]] is not None else ""
            price_raw = r[idx["price"]] if idx.get("price") is not None else None
            if not name or price_raw is None:
                continue
            try:
                price = float(str(price_raw).replace("$", "").replace(",", "").strip())
            except Exception:
                continue

            unit = None
            if "unit" in idx and idx["unit"] < len(r) and r[idx["unit"]] is not None:
                unit = str(r[idx["unit"]]).strip() or None

            vat_rate = None
            if "vat_rate" in idx and idx["vat_rate"] < len(r) and r[idx["vat_rate"]] is not None:
                try:
                    vat_rate = float(r[idx["vat_rate"]])
                except Exception:
                    vat_rate = None

            sku = None
            if "sku" in idx and idx["sku"] < len(r) and r[idx["sku"]] is not None:
                sku_val = str(r[idx["sku"]]).strip()
                sku = sku_val if sku_val else None

            parsed_rows.append({"name": name, "price": price, "unit": unit, "vat_rate": vat_rate, "sku": sku})

        def _clean_name(n):
            return re.sub(r'[\"\'\\]', '', n)

        def _batch_synonyms(names_batch: list) -> dict:
            if not openai_client:
                return {}
            try:
                numbered = "\n".join(f"{i+1}. {_clean_name(n)}" for i, n in enumerate(names_batch))
                resp = openai_client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": (
                            "Eres experto en ferreterías de México. Para cada producto numerado, "
                            "genera hasta 5 sinónimos coloquiales, typos comunes o marcas usadas como nombre genérico. "
                            'Responde SOLO en JSON válido así: {"1": "sin1, sin2", "2": "sin1, sin2"} '
                            "Sin explicación, sin markdown, solo el JSON."
                        )},
                        {"role": "user", "content": numbered}
                    ],
                    temperature=0.3, max_tokens=300,
                )
                raw = (resp.choices[0].message.content or "{}").strip()
                raw = raw.replace("```json", "").replace("```", "").strip()
                parsed = json.loads(raw)
                return {
                    names_batch[int(k)-1]: v
                    for k, v in parsed.items()
                    if k.isdigit() and int(k)-1 < len(names_batch)
                }
            except Exception as e:
                log.error("BATCH SYNONYMS ERROR: %s", repr(e))
                return {}

        synonyms_map = {}
        names_list = [r["name"] for r in parsed_rows]
        for i in range(0, len(names_list), 20):
            batch = names_list[i:i+20]
            synonyms_map.update(_batch_synonyms(batch))

        rows_total = len(parsed_rows)
        rows_upserted = 0
        rows_skipped = 0

        for row in parsed_rows:
            name = row["name"]
            name_norm = _norm_name(name)
            auto_syn = synonyms_map.get(name, "")
            try:
                cur.execute(
                    """
                    INSERT INTO pricebook_items
                        (company_id, sku, name, name_norm, unit, price, vat_rate, synonyms, source, updated_at)
                    VALUES
                        (%s, %s, %s, %s, %s, %s, %s, %s, 'excel', now())
                    ON CONFLICT (company_id, name_norm)
                    DO UPDATE SET
                        sku = EXCLUDED.sku, name = EXCLUDED.name, unit = EXCLUDED.unit,
                        price = EXCLUDED.price, vat_rate = EXCLUDED.vat_rate,
                        synonyms = COALESCE(NULLIF(pricebook_items.synonyms, ''), EXCLUDED.synonyms),
                        source = 'excel', updated_at = now()
                    """,
                    (company_id, row["sku"], name, name_norm, row["unit"], row["price"], row["vat_rate"], auto_syn),
                )
                rows_upserted += 1
            except Exception as e:
                log.error("ROW INSERT ERROR %s: %s", name, repr(e))
                rows_skipped += 1

        cur.execute(
            "UPDATE pricebook_uploads SET status='success', rows_total=%s, rows_upserted=%s, error=NULL, finished_at=now() WHERE id=%s",
            (rows_total, rows_upserted, upload_id),
        )

        if background_tasks:
            background_tasks.add_task(_rebuild_embeddings_bg, company_id)
        else:
            try:
                rebuild_embeddings_for_company(conn, company_id)
            except Exception as e:
                log.error("EMBEDDINGS REBUILD ERROR: %s", repr(e))

        return {"ok": True, "company_id": company_id, "upload_id": str(upload_id), "rows_total": rows_total, "rows_upserted": rows_upserted, "rows_skipped": rows_skipped}

    except HTTPException:
        raise
    except Exception as e:
        log.error("UPLOAD ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.get("/api/pricebook/items")
def pricebook_items(
    request: Request,
    authorization: str = Header(default=""),
    q: Optional[str] = Query(default=None),
    limit: int = Query(default=20, ge=1, le=1000),
):
    conn = None
    cur = None
    try:
        if authorization and authorization.lower().startswith("bearer "):
            company_id = get_company_from_bearer(authorization)["company_id"]
        else:
            _ = get_user_from_session(request)
            company_id = require_company_id(request)
        if not company_id:
            raise HTTPException(status_code=400, detail="No pude resolver company_id")
        conn = get_conn()
        cur = conn.cursor()
        if q:
            like = f"%{q.strip()}%"
            cur.execute(
                """
                select id, company_id, sku, name, unit, price, vat_rate, source, updated_at, created_at, bundle_size, coalesce(is_default, false)
                from pricebook_items
                where company_id = %s and (sku ilike %s or name ilike %s or name_norm ilike %s)
                order by name asc limit %s
                """,
                (company_id, like, like, like, limit),
            )
        else:
            cur.execute(
                "select id, company_id, sku, name, unit, price, vat_rate, source, updated_at, created_at, bundle_size, coalesce(is_default, false) from pricebook_items where company_id = %s order by name asc limit %s",
                (company_id, limit),
            )
        rows = cur.fetchall()
        items = []
        for r in rows:
            items.append({
                "id": r[0], "company_id": r[1], "sku": r[2],
                "name": r[3], "description": r[3], "unit": r[4],
                "price": float(r[5]) if r[5] is not None else None,
                "vat_rate": float(r[6]) if r[6] is not None else None,
                "source": r[7],
                "updated_at": r[8].isoformat() if r[8] else None,
                "created_at": r[9].isoformat() if r[9] else None,
                "bundle_size": r[10],
                "is_default": bool(r[11]) if len(r) > 11 else False,
            })
        # Get total count (independent of limit)
        cur.execute(
            "SELECT COUNT(*) FROM pricebook_items WHERE company_id=%s",
            (company_id,),
        )
        total = cur.fetchone()[0] or 0

        return {"ok": True, "items": items, "total": total}
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            try: conn.rollback()
            except Exception: pass
        log.error("PRICEBOOK ITEMS ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"pricebook_items failed: {type(e).__name__}: {e}")
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.post("/api/pricebook/items")
def pricebook_item_create(request: Request, body: PricebookItemCreateBody):
    _ = get_user_from_session(request)
    company_id = require_company_id(request)
    if not company_id:
        raise HTTPException(status_code=500, detail="DEFAULT_COMPANY_ID missing en Render")

    name = _normalize_display_name(body.name or "")
    if not name:
        raise HTTPException(status_code=400, detail="name requerido")

    sku = (body.sku or "").strip() or None
    unit = (body.unit or "").strip() or None
    source = (body.source or "manual").strip() or "manual"

    price = body.price
    if price is not None:
        try: price = float(price)
        except Exception: raise HTTPException(status_code=400, detail="price inválido")
        if price < 0: raise HTTPException(status_code=400, detail="price debe ser >= 0")

    vat_rate = body.vat_rate
    if vat_rate is not None:
        try: vat_rate = float(vat_rate)
        except Exception: raise HTTPException(status_code=400, detail="vat_rate inválido")
        if vat_rate < 0 or vat_rate > 1: raise HTTPException(status_code=400, detail="vat_rate debe estar entre 0 y 1")

    name_norm = _norm_name(name)

    # Auto-generar plurales/singulares como sinónimos
    synonyms = (body.synonyms or "").strip()
    _auto_vars = _auto_plural_singular(name)
    if _auto_vars:
        existing_set = {s.strip().lower() for s in synonyms.split(",") if s.strip()}
        new_vars = [v for v in _auto_vars if v not in existing_set]
        if new_vars:
            synonyms = (synonyms + ", " + ", ".join(new_vars)).strip(", ")

    bundle_size = body.bundle_size
    if bundle_size is not None:
        bundle_size = int(bundle_size) if bundle_size > 0 else None

    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO pricebook_items (company_id, sku, name, name_norm, unit, price, vat_rate, synonyms, source, bundle_size, updated_at, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), now())
            ON CONFLICT (company_id, name_norm)
            DO UPDATE SET sku=EXCLUDED.sku, name=EXCLUDED.name, unit=EXCLUDED.unit,
                price=EXCLUDED.price, vat_rate=EXCLUDED.vat_rate,
                synonyms=COALESCE(NULLIF(pricebook_items.synonyms,''), EXCLUDED.synonyms),
                source=EXCLUDED.source, bundle_size=EXCLUDED.bundle_size, updated_at=now()
            RETURNING id
            """,
            (company_id, sku, name, name_norm, unit, price, vat_rate, synonyms, source, bundle_size),
        )
        new_id = cur.fetchone()[0]
        try:
            upsert_single_embedding(conn, company_id, new_id, name, sku or "", unit or "", synonyms or "")
        except Exception as e:
            log.error("SINGLE EMBEDDING ERROR: %s", repr(e))
        return {"ok": True, "id": str(new_id)}
    except IntegrityError as e:
        msg = str(e).lower()
        if "duplicate" in msg or "unique" in msg:
            raise HTTPException(status_code=409, detail="Producto ya existe (conflicto)")
        raise HTTPException(status_code=400, detail="Integridad inválida")
    except HTTPException:
        raise
    except Exception as e:
        log.error("PRICEBOOK CREATE ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="pricebook_item_create failed")
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.post("/api/pricebook/bulk")
def pricebook_bulk_create(request: Request, body: PricebookBulkBody):
    """Bulk create products (used by onboarding wizard)."""
    company_id = require_company_id(request)
    if not body.items:
        raise HTTPException(status_code=400, detail="items vacío")
    if len(body.items) > 50:
        raise HTTPException(status_code=400, detail="Máximo 50 productos por lote")

    conn = None
    cur = None
    created = 0
    try:
        conn = get_conn()
        cur = conn.cursor()
        for item in body.items:
            name = _normalize_display_name((item.get("name") or "").strip())
            if not name:
                continue
            price = item.get("price")
            try:
                price = float(price) if price is not None else None
            except (ValueError, TypeError):
                price = None
            unit = (item.get("unit") or "Pieza").strip()
            name_n = _norm_name(name)

            # Auto-generate plural/singular synonyms
            synonyms_list = _auto_plural_singular(name)
            synonyms = ", ".join(synonyms_list) if synonyms_list else ""

            cur.execute(
                """
                INSERT INTO pricebook_items (company_id, name, name_norm, unit, price, vat_rate, synonyms, source, updated_at, created_at)
                VALUES (%s, %s, %s, %s, %s, 0.16, %s, 'onboarding', now(), now())
                ON CONFLICT (company_id, name_norm)
                DO UPDATE SET price = EXCLUDED.price, unit = EXCLUDED.unit, updated_at = now()
                RETURNING id
                """,
                (company_id, name, name_n, unit, price, synonyms),
            )
            row = cur.fetchone()
            if row:
                created += 1
                try:
                    upsert_single_embedding(conn, company_id, row[0], name, "", unit, synonyms)
                except Exception as e:
                    log.error("BULK EMBEDDING ERROR for '%s': %s", name, repr(e))
        conn.commit()
        # Auto-generate context groups after bulk product upload
        if created >= 3:
            try:
                cg_result = auto_generate_context_groups(conn, company_id)
                log.debug("BULK CONTEXT GROUPS: %s", cg_result.get("status"))
            except Exception as cge:
                log.error("BULK CONTEXT GROUPS ERROR: %s", repr(cge))
        return {"ok": True, "created": created}
    except HTTPException:
        raise
    except Exception as e:
        log.error("PRICEBOOK BULK ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error en carga masiva")
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.delete("/api/pricebook/items/{item_id}")
def pricebook_item_delete(request: Request, item_id: str):
    _ = get_user_from_session(request)
    company_id = require_company_id(request)
    if not company_id:
        raise HTTPException(status_code=500, detail="DEFAULT_COMPANY_ID missing en Render")
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM pricebook_items WHERE company_id = %s AND id = %s RETURNING id", (company_id, item_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        log.error("PRICEBOOK DELETE ERROR: %s", repr(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="pricebook_item_delete failed")
    finally:
        if cur: cur.close()
        if conn: conn.close()


@router.patch("/api/pricebook/items/{item_id}")
def pricebook_item_update(request: Request, item_id: str, body: PricebookItemUpdateBody):
    _ = get_user_from_session(request)
    company_id = require_company_id(request)

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT name, sku, unit, price, vat_rate, synonyms, bundle_size, coalesce(is_default, false) FROM pricebook_items WHERE id=%s AND company_id=%s",
            (item_id, company_id),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Producto no encontrado")

        name     = (body.name.strip() if body.name is not None else None) or row[0]
        sku      = (body.sku.strip()  if body.sku  is not None else None) or row[1]
        unit     = (body.unit.strip() if body.unit is not None else None) or row[2]
        price    = body.price    if body.price    is not None else row[3]
        vat_rate = body.vat_rate if body.vat_rate is not None else row[4]
        synonyms = (body.synonyms.strip() if body.synonyms is not None else None) or row[5] or ""
        bundle_size = body.bundle_size if body.bundle_size is not None else row[6]
        if bundle_size is not None and bundle_size <= 0:
            bundle_size = None
        is_default = body.is_default if body.is_default is not None else row[7]

        # Auto-generar plurales/singulares como sinónimos
        _auto_vars = _auto_plural_singular(name)
        if _auto_vars:
            existing_set = {s.strip().lower() for s in synonyms.split(",") if s.strip()}
            new_vars = [v for v in _auto_vars if v not in existing_set]
            if new_vars:
                synonyms = (synonyms + ", " + ", ".join(new_vars)).strip(", ")

        name_norm = _norm_name(name)

        cur.execute(
            """
            UPDATE pricebook_items
            SET name=%s, name_norm=%s, sku=%s, unit=%s, price=%s, vat_rate=%s, synonyms=%s, bundle_size=%s, is_default=%s, updated_at=now()
            WHERE id=%s AND company_id=%s
            RETURNING id
            """,
            (name, name_norm, sku, unit, price, vat_rate, synonyms, bundle_size, is_default, item_id, company_id),
        )
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Producto no encontrado")

        try:
            upsert_single_embedding(conn, company_id, item_id, name, sku or "", unit or "", synonyms or "")
        except Exception as e:
            log.error("EMBEDDING UPDATE ERROR: %s", repr(e))

        return {"ok": True}
    finally:
        cur.close()
        conn.close()


@router.get("/api/pricebook/items/{item_id}/synonyms-suggestions")
def synonyms_suggestions(request: Request, item_id: str):
    _ = get_user_from_session(request)
    company_id = require_company_id(request)
    openai_client = _get_openai_client()
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT name, synonyms FROM pricebook_items WHERE id=%s AND company_id=%s", (item_id, company_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        name = (row[0] or "").strip()
        existing = (row[1] or "")
        suggestions = []
        if openai_client and name:
            try:
                resp = openai_client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": "Eres un experto en ferreterías de México. Dado un producto, devuelve palabras alternativas coloquiales que usarían clientes o ferreteros para pedirlo. Responde SOLO con las palabras separadas por coma, sin explicación, sin puntos, en minúsculas."},
                        {"role": "user", "content": f"Producto: {name}\nDame 4 sinónimos o nombres alternativos coloquiales."}
                    ],
                    temperature=0.3, max_tokens=60,
                )
                raw = resp.choices[0].message.content or ""
                suggestions = [s.strip().lower() for s in raw.split(",") if s.strip()]
            except Exception as e:
                log.error("SYNONYMS GPT ERROR: %s", repr(e))
        existing_list = [s.strip().lower() for s in existing.split(",") if s.strip()]
        suggestions = [s for s in suggestions if s not in existing_list]
        return {"ok": True, "suggestions": suggestions, "existing": existing_list}
    finally:
        cur.close()
        conn.close()


@router.post("/api/pricebook/deduplicate")
def pricebook_deduplicate(request: Request):
    company_id = require_company_id(request)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            DELETE FROM pricebook_items
            WHERE id NOT IN (
                SELECT MIN(id::text)::uuid FROM pricebook_items WHERE company_id=%s GROUP BY name_norm
            ) AND company_id=%s
            """,
            (company_id, company_id),
        )
        deleted = cur.rowcount
        return {"ok": True, "deleted": deleted}
    finally:
        cur.close()
        conn.close()
